# Copyright (c) 2026 Beijing Volcano Engine Technology Co., Ltd.
# SPDX-License-Identifier: Apache-2.0
"""
Excel (.xlsx/.xls/.xlsm) parser for OpenViking.

Converts Excel spreadsheets to Markdown then parses using MarkdownParser.
Inspired by microsoft/markitdown approach.
"""

from pathlib import Path
from typing import List, Optional, Union

from openviking.parse.base import ParseResult
from openviking.parse.parsers.base_parser import BaseParser
from openviking_cli.utils.config.parser_config import ParserConfig
from openviking_cli.utils.logger import get_logger

logger = get_logger(__name__)


class ExcelParser(BaseParser):
    """
    Excel spreadsheet parser for OpenViking.

    Supports: .xlsx, .xls, .xlsm

    Converts Excel spreadsheets to Markdown using openpyxl,
    then delegates to MarkdownParser for tree structure creation.
    """

    def __init__(self, config: Optional[ParserConfig] = None, max_rows_per_sheet: int = 1000):
        """
        Initialize Excel parser.

        Args:
            config: Parser configuration
            max_rows_per_sheet: Maximum rows to process per sheet (0 = unlimited)
        """
        from openviking.parse.parsers.markdown import MarkdownParser

        self._md_parser = MarkdownParser(config=config)
        self.config = config or ParserConfig()
        self.max_rows_per_sheet = max_rows_per_sheet

    @property
    def supported_extensions(self) -> List[str]:
        return [".xlsx", ".xls", ".xlsm"]

    async def parse(self, source: Union[str, Path], instruction: str = "", **kwargs) -> ParseResult:
        """Parse Excel spreadsheet from file path."""
        path = Path(source)

        if path.exists():
            import openpyxl

            markdown_content = self._convert_to_markdown(path, openpyxl)
            result = await self._md_parser.parse_content(
                markdown_content, source_path=str(path), instruction=instruction, **kwargs
            )
        else:
            result = await self._md_parser.parse_content(
                str(source), instruction=instruction, **kwargs
            )
        result.source_format = "xlsx"
        result.parser_name = "ExcelParser"
        return result

    async def parse_content(
        self, content: str, source_path: Optional[str] = None, instruction: str = "", **kwargs
    ) -> ParseResult:
        """Parse content - delegates to MarkdownParser."""
        result = await self._md_parser.parse_content(content, source_path, **kwargs)
        result.source_format = "xlsx"
        result.parser_name = "ExcelParser"
        return result

    def _convert_to_markdown(self, path: Path, openpyxl) -> str:
        """Convert Excel spreadsheet to Markdown string."""
        wb = openpyxl.load_workbook(path, data_only=True)

        markdown_parts = []
        markdown_parts.append(f"# {path.stem}")
        markdown_parts.append(f"**Sheets:** {len(wb.sheetnames)}")

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_content = self._convert_sheet(sheet, sheet_name)
            markdown_parts.append(sheet_content)

        return "\n\n".join(markdown_parts)

    def _convert_sheet(self, sheet, sheet_name: str) -> str:
        """Convert a single sheet to markdown."""
        parts = []
        parts.append(f"## Sheet: {sheet_name}")

        max_row = sheet.max_row
        max_col = sheet.max_column

        if max_row == 0 or max_col == 0:
            parts.append("*Empty sheet*")
            return "\n\n".join(parts)

        parts.append(f"**Dimensions:** {max_row} rows × {max_col} columns")

        rows_to_process = max_row
        if self.max_rows_per_sheet > 0:
            rows_to_process = min(max_row, self.max_rows_per_sheet)

        rows = []
        for _row_idx, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=rows_to_process, values_only=True), 1
        ):
            row_data = []
            for cell in row:
                if cell is None:
                    row_data.append("")
                else:
                    row_data.append(str(cell))
            rows.append(row_data)

        if rows:
            from openviking.parse.base import format_table_to_markdown

            table_md = format_table_to_markdown(rows, has_header=True)
            parts.append(table_md)

        if self.max_rows_per_sheet > 0 and max_row > self.max_rows_per_sheet:
            parts.append(f"\n*... {max_row - self.max_rows_per_sheet} more rows truncated ...*")

        return "\n\n".join(parts)
